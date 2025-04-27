# Import required libraries
import pandas as pd  # pandas for reading/writing Excel and data manipulation
import networkx as nx  # networkx for graph operations (optional here)
from random import shuffle  # shuffle for randomizing lists

# Load students and their chosen friends from an Excel file
def load_students(filename, sheet_name="students"):
    df = pd.read_excel(filename, sheet_name=sheet_name)  # Read the Excel file
    students = []  # List to store all student names
    student_to_friends = {}  # Dictionary to map each student -> list of chosen friends

    for _, row in df.iterrows():  # Iterate over each row in the file
        student = row["student"]  # Get student name
        friends = []  # Initialize list of this student's chosen friends

        # If friend1 exists (not NaN), add it to the friends list
        if pd.notna(row.get("friend1")):
            friends.append(row["friend1"])

        # If friend2 exists (not NaN), add it to the friends list
        if pd.notna(row.get("friend2")):
            friends.append(row["friend2"])

        students.append(student)  # Add student to the student list
        student_to_friends[student] = friends  # Map student to their friends

    return students, student_to_friends  # Return the two structures

# Build an undirected graph where students are nodes and friendships are edges
def build_friend_graph(students, student_to_friends):
    G = nx.Graph()  # Initialize an empty undirected graph

    for student in students:  # For each student
        G.add_node(student)  # Add student as a node

        for friend in student_to_friends.get(student, []):  # For each friend they chose
            G.add_edge(student, friend)  # Add an undirected edge between student and friend

    return G  # Return the constructed graph

# Assign students to groups, maximizing friends together
def assign_students_to_groups(students, student_to_friends, num_groups):
    groups = {f"Group{i+1}": [] for i in range(num_groups)}  # Initialize groups dictionary
    student_to_group = {}  # Track which student is assigned to which group

    shuffle(students)  # Shuffle students to randomize order

    for student in students:  # Go through each student
        best_group = None  # Keep track of best group for this student
        best_score = -float('inf')  # Initialize best score as very negative

        for group, members in groups.items():  # Check all groups
            if student in student_to_group:  # If already assigned, skip
                continue

            # Calculate how many of this student's friends are in the group
            friend_score = sum(1 for f in student_to_friends.get(student, []) if f in members)

            # Penalize big groups to prefer balanced group sizes
            size_penalty = len(members)

            # Final score: more friends = better, bigger group = worse
            score = friend_score * 10 - size_penalty

            # If this group is the best so far, remember it
            if score > best_score:
                best_score = score
                best_group = group

        # After checking all groups, assign student to best group
        if best_group is not None:
            groups[best_group].append(student)  # Add student to group
            student_to_group[student] = best_group  # Record the assignment

    return groups  # Return final groups

# Try to balance the sizes of the groups
def balance_groups(groups, target_size):
    group_list = list(groups.keys())  # List of group names

    while True:  # Keep balancing until no more changes
        # Groups that are too big
        overfilled = [g for g in group_list if len(groups[g]) > target_size + 1]

        # Groups that are too small
        underfilled = [g for g in group_list if len(groups[g]) < target_size]

        # If no group needs adjustment, stop
        if not overfilled or not underfilled:
            break

        donor = overfilled[0]  # Take from the first overfilled group
        receiver = underfilled[0]  # Add to the first underfilled group
        student = groups[donor].pop()  # Remove one student from donor
        groups[receiver].append(student)  # Add student to receiver

    return groups  # Return balanced groups

# Evaluate how many students were assigned together with at least one friend
def evaluate_friend_success(groups, student_to_friends):
    # Map each student to their group
    student_to_group = {
        student: group
        for group, members in groups.items()
        for student in members
    }

    total = 0  # Total number of students evaluated
    success = 0  # Number of successful friend assignments

    for student, friends in student_to_friends.items():
        if student not in student_to_group:  # Ignore students not assigned
            continue

        group = student_to_group[student]  # Find student's group

        # Check if at least one friend is in the same group
        if any(f in student_to_group and student_to_group[f] == group for f in friends):
            success += 1  # Friend found in same group -> success

        total += 1  # Count this student

    # Calculate success percentage
    percent = (success / total) * 100 if total > 0 else 0

    return percent, success, total  # Return results

# Export the groups to an Excel file, with coloring for each group
def export_assignments_to_excel(groups, student_to_friends, excel_file, sheet_name="final_assignment"):
    from openpyxl import load_workbook  # Open existing Excel file
    from openpyxl.styles import PatternFill  # Style cells

    # List of colors to cycle through
    colors = [
        "FFC7CE", "C6EFCE", "FFEB9C", "DDEBF7",
        "F8CBAD", "E2EFDA", "D9E1F2", "FCE4D6", "F4CCCC"
    ]

    rows = []  # List of rows to export

    # Prepare the data for export
    for group, members in groups.items():
        for student in members:
            friends = student_to_friends.get(student, [])
            friends_in_group = [f for f in friends if f in members]
            rows.append({
                "Group": group,
                "Student": student,
                "Friends Chosen": ", ".join(friends),
                "Friends in Same Group": ", ".join(friends_in_group),
                "Success": "✔️" if friends_in_group else "❌"
            })

    df = pd.DataFrame(rows)  # Create DataFrame from rows
    df = df.sort_values(by=["Group", "Student"])  # Sort nicely

    try:
        # Try to open existing Excel and overwrite the sheet
        with pd.ExcelWriter(excel_file, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)

        wb = load_workbook(excel_file)  # Load the workbook
        ws = wb[sheet_name]  # Access the new sheet

        # Map groups to colors
        unique_groups = sorted(df["Group"].unique())
        group_color_map = {
            group: PatternFill(start_color=colors[i % len(colors)], end_color=colors[i % len(colors)], fill_type="solid")
            for i, group in enumerate(unique_groups)
        }

        # Apply colors row by row
        for row_idx in range(2, len(df) + 2):
            group_value = ws[f"A{row_idx}"].value
            fill = group_color_map.get(group_value)
            if fill:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col).fill = fill

        wb.save(excel_file)  # Save changes
        print(f"\nExported to Excel with coloring in sheet: '{sheet_name}'")

    except FileNotFoundError:
        # If Excel file doesn't exist, create new one
        df.to_excel(excel_file, index=False, sheet_name=sheet_name)
        print(f"\nSaved new Excel file with sheet: '{sheet_name}'")

# Complete process: Load, assign, balance, and evaluate students
def run_smart_assignment(filename, num_groups=9):
    students, student_to_friends = load_students(filename)  # Load data
    shuffle(students)  # Shuffle students

    groups = assign_students_to_groups(students, student_to_friends, num_groups)  # Assign groups

    total_students = len(students)  # Total number of students
    target_size = total_students // num_groups  # Target group size
    groups = balance_groups(groups, target_size=target_size)  # Balance groups

    percent, success, total = evaluate_friend_success(groups, student_to_friends)  # Evaluate success

    return percent, groups, student_to_friends, success, total  # Return all results

# Print group assignments
def print_final_groups(groups):
    print("\nFinal group assignments:")
    for group, members in groups.items():
        print(f"\n- {group} – {len(members)} students:")
        for student in sorted(members):
            print(f"  - {student}")

# Print the friend success percentage
def print_percent(percent, success, total):
    print(f"\nFriend Success Rate: {success}/{total} students ({percent:.2f}%)")

# Main function to run the assignment process multiple times and keep the best result
def run_n_times(num_groups=9,n=50):
    filename = "students_table_new.xlsx"  # Output file name

    max_values_dict = {"total": 0, "success": 0, "percent": 0, "groups": None, "student_to_friends": None}
    curr_values_dict = {"total": 0, "success": 0, "percent": 0, "groups": None, "student_to_friends": None}

    for i in range(n):  # Repeat assignment n times
        curr_values_dict["percent"], curr_values_dict["groups"], curr_values_dict["student_to_friends"], curr_values_dict["success"], curr_values_dict["total"] = run_smart_assignment(filename, num_groups)

        # If current success rate is higher, update best values
        if curr_values_dict["percent"] > max_values_dict["percent"]:
            max_values_dict["percent"] = curr_values_dict["percent"]
            max_values_dict["groups"] = curr_values_dict["groups"]
            max_values_dict["student_to_friends"] = curr_values_dict["student_to_friends"]
            max_values_dict["total"] = curr_values_dict["total"]
            max_values_dict["success"] = curr_values_dict["success"]

    # Export the best groups to Excel
    export_assignments_to_excel(max_values_dict["groups"], max_values_dict["student_to_friends"], excel_file=filename)

    # Print final groups and stats
    print_final_groups(max_values_dict["groups"])
    print_percent(max_values_dict["percent"], max_values_dict["success"], max_values_dict["total"])

# Actually run the function for 500 iterations
run_n_times()
